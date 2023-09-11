#     ______                                         __
#    / ____/____   _____ _      __ ____ _ _____ ____/ /
#   / /_   / __ \ / ___/| | /| / // __ `// ___// __  /
#  / __/  / /_/ // /    | |/ |/ // /_/ // /   / /_/ /
# /_/ _   \____//_/__   |__/|__/ \__,_//_/    \__,_/
#    / | / /___   / /_ _      __ ____   _____ / /__ _____
#   /  |/ // _ \ / __/| | /| / // __ \ / ___// //_// ___/
#  / /|  //  __// /_  | |/ |/ // /_/ // /   / ,<  (__  )
# /_/ |_/ \___/ \__/  |__/|__/ \____//_/   /_/|_|/____/

# Discalimer:

# This software is provided as is, without any warranty or support. Use of the software is at your own risk.
# The author and any contributors will not be held responsible for any damages or issues that may arise from
# the use of this software.

# Please be aware that this software may contain bugs, errors, or other issues.
# It may not function as intended or be fit for your specific use case.

# By using this software, you acknowledge and accept the above disclaimer and assume
# all responsibility for any issues that may arise from its use.

# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

# Copyright 2023 Forward Networks, Inc.
# All rights reserved.


"""
Usage:
  fwd-intent-gen.py from_import <input> <appserver> <snapshot> [--batch=<batch_size>] [--limit=<limit>] [--max=<max_query>] [--withdiag] [--debug]
  fwd-intent-gen.py from_acls <appserver> <snapshot>   [--batch=<batch_size>] [--limit=<limit>] [--max=<max_query>] [--withdiag] [--debug]
  fwd-intent-gen.py check <input> <appserver> <snapshot> [--csv] [--debug]

Options:
  -h --help             Show this help message
  --batch=<batch_size>  Configure batch size [default: 300]
  --csv                 "Dump into CSV file for import"
  --debug               "Set Debug Flag [default: False]"
  --limit=<limit>       "Limit to n applications (ACL-names) [default: 1000]
  --max=<max_query>    "Max queries [default: 10000]
"""

import itertools
import math
import re
import socket
import sys
import traceback
import pandas as pd
import aiohttp
import asyncio
import json
import os
from docopt import docopt

# from openpyxl.styles import Font
from openpyxl import load_workbook
import requests
import logging
import glob
import datetime
from tqdm import tqdm
from urllib.parse import quote
import urllib3
import inspect
from ipaddress import ip_network, ip_address


urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


forwardingOutcomes = {
    "DELIVERED": {
        "description": "Traffic was delivered to destination IP’s discovered location(s).",
        "remedy": "None",
    },
    "NOT_DELIVERED": {
        "description": "No available paths found for query",
        "remedy": "This error may indicate that the source or destination address could not be found in the model. Check the source and destination address using Forward Enterprise search. The address may be incorrect or the address cannot be located.",
    },
    "DELIVERED_TO_INCORRECT_LOCATION": {
        "description": "Traffic was delivered out of some edge ports. However, traffic did not reach the expected delivery locations based on destination IP’s discovered locations. One scenario where this occurs is when a device in the middle of the actual path from source IP to destination IP is not configured for collection in the Forward platform. For example, suppose the actual device path is A -> B -> C, and only devices A and C are part of the snapshot in the Forward platform. In this case, the path would show traffic exiting device A at some edge port, but since destination IP is discovered to reside at device C, traffic is delivered to an incorrect location.",
        "remedy": " This error should not happen on a fully modeled network, the result_df is indicating there is a high chance there is a missing device in the Forward Enteprise Model. Leverage the pathsURL to diagnose where the last hop device and interface is, Forward Enterprise will report any missing devices as indicated by CDP/LLDP. Work with your teams to assess what device is missing and add to the model",
    },
    "BLACKHOLE": {
        "description": "Traffic was implicitly dropped at the last hop, since the device had no matching rule. For example, if a router does not have a default route, traffic to any IP that is not in the routing table gets blackholed.",
        "remedy": "This error should not typically happen on a production network, it indicates that there is missing information for the device to forward traffic to another device. Leverage the queryURL to investigate using the Forward Enterprise platform to understand why there is no next-hop to forward traffic to.",
    },
    "DROPPED": {
        "description": "Traffic was explicitly dropped at the last hop, e.g. by a null route.",
        "remedy": "Check the last hop device to ensure there is a route to the destination",
    },
    "INADMISSIBLE": {
        "description": "Traffic was not admitted into the network. The first hop interface does not accept the traffic, e.g. incoming traffic had a vlan tag 10 while the ingress interface is an access interface that only permits traffic with vlan tag 20.",
        "remedy": "This error may indicate an incorrect assumption about the where to originate the source of the check.",
    },
    "UNREACHABLE": {
        "description": "ARP/NDP resolution failed along the path resulting in traffic not getting delivered to the intended destination.",
        "remedy": "This error may indicate missing state in the model or state that has aged out due to inactivity. Use the queryURL to identify the next-hop and identify if that device is missing necessary reachability information",
    },
    "LOOP": {
        "description": "Traffic entered a forwarding loop.",
        "remedy": "This error should be rare but can happen if improper static routes are defined with a default gateway. Use the queryURL to examine the path to determine which device has the incorrect route configuration",
    },
}
securityOutcomes = {
    "PERMITTED": {
        "description": "All ACLs along the path permitted traffic to flow through.",
        "remedy": "None",
    },
    "DENIED": {
        "description": "Traffic was dropped by ACLs at some hop along the path. Note that the ACL drop may not always occur at the last hop since search results are computed in permit all mode.",
        "remedy": "Traffic is denied by a security policy. Update the ACL or firewall policy to allow communications and retest",
    },
    "UNKNOWN": {
        "description": "No path has been found for this search so a security outcome is reported as UNKNOWN",
        "remedy": "This error may indicate that the source or destination address could not be found in the model. Check the source and destination address using Forward Enterprise search. The address may be incorrect or the address cannot be located.",
    },
}

acl_query = """
getAcl =
  foreach device in network.devices
  foreach aclEntry in device.aclEntries
  select {
    sources: (foreach s in aclEntry.headerMatches.ipv4Src
              select s),
    destinations: (foreach d in aclEntry.headerMatches.ipv4Dst
                   select d),
    action:    when aclEntry.action is
                       DENY -> "DENY";
                       PBR -> "PBR";
                       PERMIT -> "PERMIT",
    protocols: (foreach t in aclEntry.headerMatches.ipProtocol select {start: t.start, end: t.end}),
    dstports: (foreach p in aclEntry.headerMatches.tpDst select {start: p.start, end: p.end}),
    name: aclEntry.name,
    device: device.name
  };

getHosts =
  foreach d in network.devices
  foreach hosts in d.hosts
  foreach host in hosts.addresses
  select host;

foreach x in [0]
let acls = distinct(getAcl())
let hosts = getHosts()
foreach acl in acls
foreach host in hosts
where host in acl.sources || host in acl.destinations
group acl as a
  by { name: acl.name,
       src: acl.sources,
       dst: acl.destinations,
       protos: acl.protocols,
       dstPorts: acl.dstports,
       hostCount: length(hosts),
       action: acl.action,
       device: acl.device
    }
    as b
select distinct { application: b.name, sources: b.src, destinations: b.dst, protocols: b.protos, dstPorts: b.dstPorts, action: b.action, hostcount: b.hostCount, device: b.device}
"""

host_query = """ 
foreach device in network.devices
foreach host in device.hosts
foreach hostSubnet in host.addresses
foreach interface in host.interfaces
where host.hostType == DeviceHostType.InferredHost
select {
  deviceName: device.name,
  Address: address(hostSubnet),
  MacAddress: host.macAddress,
  OUI: if isPresent(host.macAddress) then ouiAssignee(host.macAddress) else "",
  HostType: host.hostType,
  Interface: interface,
}
"""

options = {
    "intent": "PREFER_DELIVERED",
    "maxCandidates": 5000,
    "maxResults": 1,
    "maxReturnPathResults": 1,
    "maxSeconds": 30,
    "maxOverallSeconds": 60,
    "includeNetworkFunctions": False,
}

headers_seq = {
    "Accept": "application/json-seq",
    "Content-Type": "application/json",
}

headers = {
    "Accept": "application/json",
    "Content-Type": "application/json",
}

# Utilities


def getDisposition(
    dstIp,
    hosts,
    egressInterface,
    dstIpLocationType,
    securityOutcome,
    forwardingOutcome,
    dest_status,
):
    if (
        forwardingOutcome == "DELIVERED_TO_INCORRECT_LOCATION"
        and dest_status == "VALID"
    ):
        return "INSUFFICIENT_INFO"
    elif hosts is not None and any(
        ip_network(dstIp).network_address in ip_network(address) for address in hosts
    ):
        return "ACCEPTED"
    elif (
        dstIpLocationType == "INTERFACE_ATTACHED_SUBNET"
        and forwardingOutcome == "DELIVERED"
        and securityOutcome == "PERMITTED"
    ):
        return "DELIVERED_TO_SUBNET"
    elif (
        egressInterface == "self"
        and forwardingOutcome == "DELIVERED"
        and securityOutcome == "PERMITTED"
        and dstIpLocationType == "INTERFACE"
    ):
        return "ACCEPTED"
    else:
        return "UNKNOWN"


def getDiagnostic(
    dstIp,
    hosts,
    egressInterface,
    dstIpLocationType,
    securityOutcome,
    forwardingOutcome,
    dest_status,
    behaviors,
    app_df,
    violation,
    aclAction
):

    shadowed = False
    for behavior in behaviors:
        if behavior[1] == "ACL_DENY" and violation == True:
            shadowed = True
            break
    if shadowed:
        return "SHADOWED_ACL"
    else:
        return "NONE"


def print_debug(message):
    print(
        f"{datetime.datetime.now()} Debug on line {inspect.currentframe().f_back.f_lineno}: {message}"
    )


def toQuote(s):
    return quote(s, safe="")


def test_communication(appserver):
    try:
        response = requests.get(f"https://{appserver}", timeout=10, verify=False)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f"HTTP communication failed with {appserver} {e}.")
        sys.exit(1)


def flatten_input(data):
    rows = []
    for region, apps in data.items():
        for app, details in apps.items():
            rows.append(
                [
                    region,
                    app,
                    details["source"],
                    details["destination"],
                    details["ipProto"],
                    details["dstPorts"],
                    details["action"],
                ]
            )
    return pd.DataFrame(
        rows,
        columns=[
            "region",
            "application",
            "sources",
            "destinations",
            "protocols",
            "dstPorts",
            "action",
        ],
    )


def remove_columns_df(df, columns):
    return df.drop(columns, axis=1)


# def update_font(f):
#     workbook = load_workbook(f)
#     worksheet = workbook.active
#     font = Font(size=14)  # Set font size to 14
#     for row in worksheet.iter_rows():
#         for cell in row:
#             cell.font = font
#     workbook.save(f)


def remove_columns(data, columns_to_remove):
    new_data = []
    for item in data:
        new_item = item.copy()
        for column in columns_to_remove:
            new_item.pop(column)
        new_data.append(new_item)
    return new_data


def parse_start_end(s):
    pattern = r"start:(\d+), end:(\d+)"
    match = re.search(pattern, s)
    if match:
        start_value = str(match.group(1))
        end_value = str(match.group(2))
        if start_value == end_value:
            return start_value
        else:
            return f"{start_value}-{end_value}"


def resolve_ip_to_domain(ip_address):
    try:
        domain_name = socket.gethostbyaddr(ip_address)[0]
        return domain_name
    except socket.herror as e:
        return ip_address


# API


def nqe_get_hosts_from_acl(query, appserver, snapshot):
    url = f"https://{appserver}/api/nqe?snapshotId={snapshot}"
    limit = 10000
    offset = 0
    items = []
    total = 0

    if debug:
        print_debug("calling nqe_get_hosts_from_acl")

    with tqdm(total=total, dynamic_ncols=True) as pbar:
        while True:
            body = {"query": query, "queryOptions": {"limit": limit, "offset": offset}}
            if debug:
                print_debug(f"{body}, {url}, {headers}, {username}")
            try:
                response = requests.post(
                    url,
                    json=body,
                    auth=(username, password),
                    headers=headers,
                    verify=False,
                )
                response_status = response.status_code

                response.raise_for_status()
                response_json = response.json()
                if debug:
                    print_debug(
                        f"Offset: {offset} Items: {len(response_json['items'])} List: {len(items)} Data: {response_json}"
                    )
                if not response_json["items"]:
                    break
                items.extend(response_json["items"])
                offset += limit
                pbar.total = offset
                pbar.update(limit)
            except requests.exceptions.HTTPError as err:
                if response_status in [401, 403]:
                    print_debug(
                        "Please set FWD_USER and FWD_PASSWORD to authentication credentials"
                    )
                elif response_status in [409]:
                    print_debug("Snapshot is being processed, try again in a few")
                raise SystemExit(err)
            except requests.exceptions.ConnectionError as err:
                print_debug(f"Connection error occurred: {err}")
                break

    return items


# Get the username and password from environment variables.
username = os.getenv("FWD_USER")
password = os.getenv("FWD_PASSWORD")

# Set Debug if needed
debug = True if os.getenv("DEBUG") else False

if not username or not password:
    print("Please provide both FWD_USER and FWD_PASSWORD.")
    sys.exit()


def return_firstlast_hop(df):
    new_df = df.copy()
    # Extract values using apply()
    (
        new_df["firstHopDevice"],
        new_df["firstHopDeviceType"],
        new_df["lastHopDevice"],
        new_df["lastHopDeviceType"],
        new_df["lastHopEgressIntf"],
        new_df["ACLS"],
    ) = zip(
        *new_df["hops"].apply(
            lambda hops: (
                hops[0]["deviceName"],
                hops[0]["deviceType"],
                hops[-1]["deviceName"],
                hops[-1]["deviceType"],
                hops[-1].get("egressInterface"),
                [
                    (hop["deviceName"], behavior)
                    for hop in hops
                    if "behaviors" in hop
                    for behavior in hop["behaviors"]
                    if behavior in ["ACL_PERMIT", "ACL_DENY"]
                ],
            )
            if isinstance(hops, list) and len(hops) > 0
            else (None, None, None, None, None, [])
        )
    )
    # Remove the "hops" column
    new_df = new_df.drop(columns=["hops"])

    return new_df


def addForwardingOutcomes(result):
    result["forwardingDescription"] = result["forwardingOutcome"].apply(
        lambda x: forwardingOutcomes[x]["description"]
        if x in forwardingOutcomes
        else None
    )
    result["forwardingRemedy"] = result["forwardingOutcome"].apply(
        lambda x: forwardingOutcomes[x]["remedy"] if x in forwardingOutcomes else None
    )
    result["securityDescription"] = result["securityOutcome"].apply(
        lambda x: securityOutcomes[x]["description"] if x in securityOutcomes else None
    )
    result["securityRemedy"] = result["securityOutcome"].apply(
        lambda x: securityOutcomes[x]["remedy"] if x in securityOutcomes else None
    )

    return result


def check_info_paths(data):
    for element in data:
        info = element.get("info", {})
        paths = info.get("paths", [])
        element["pathCount"] = len(paths)
        timedOut = element.get("timedOut")
        if timedOut:
            raise Exception("Timed out")
        srcIpLocationType = element.get("srcIpLocationType", "UNKNOWN")
        dstIpLocationType = element.get("dstIpLocationType", "UNKNOWN")

        paths = paths or [
            {
                "forwardingOutcome": "NOT_DELIVERED",
                "securityOutcome": "UNKNOWN",
                "hops": [],
            }
        ]

        element["forwardHops"] = len(paths[0].get("hops", []))
        element["info"] = {"paths": paths}
        element["srcIpLocationType"] = srcIpLocationType
        element["dstIpLocationType"] = dstIpLocationType

        return_path_info = element.get("returnPathInfo", {})
        return_paths = return_path_info.get("paths", [])

        if return_paths:
            element["returnHops"] = len(return_paths[0].get("hops", []))
            element["returnPathCount"] = len(return_paths)
        else:
            element["returnPathCount"] = 0
            element["returnHops"] = 0
            return_paths = [
                {
                    "forwardingOutcome": "NOT_DELIVERED",
                    "securityOutcome": "UNKNOWN",
                    "hops": [],
                }
            ]
        element["returnPathInfo"] = {"paths": return_paths}

    return data


def parse_subnets(data):
    origin_descriptions = {
        "HOST": {
            "description": "Host Interface",
            "data_key": "hosts",
        },
        "INTERFACE": {"description": "Device Interface", "data_key": "interfaces"},
        "INTERFACE_ATTACHED_SUBNET": {
            "description": "Packet delivered to subnet",
            "data_key": "locations",
        },
        "ROUTE": {"description": "Route", "data_key": "locations"},
        "UNKNOWN": {"description": "UNKNOWN", "data_key": None},
    }

    parsed_data = [
        {
            "address": item["address"],
            "origin": item.get("origin", "UNKOWN"),
            "description": origin_descriptions.get(item.get("origin", "ERROR"), {}).get(
                "description", "INVALID"
            ),
            "status": "VALID",
            # if item.get("origin") in ["HOST", "INTERFACE"]
            # else "INVALID"
            # if item.get("origin") in origin_descriptions
            # else "INVALID",
            "data": item.get(
                origin_descriptions.get(item.get("origin", {}), {}).get("data_key")
            ),
        }
        if "origin" in item and "address" in item
        else {
            "address": item.get("address"),
            "origin": "UNKOWN",
            "description": "NOTFOUND",
            "status": "INVALID",
            "data": None,
        }
        for item in data
    ]
    if debug:
        for item in data:
            print_debug(item)
        for item in parsed_data:
            print_debug(item)

    return parsed_data


async def fetch(
    session,
    url,
    data=None,
    method="GET",
    username=None,
    password=None,
    headers={},
    timeout=60,
    retries=3,
):
    auth = aiohttp.BasicAuth(username, password) if username and password else None
    for retry in range(retries):
        try:
            if method == "GET":
                async with session.get(
                    url,
                    auth=auth,
                    params=data,
                    headers=headers,
                    ssl=False,
                    timeout=timeout,
                ) as response:
                    # print(f"GET request to {url} returned status {response.status}")
                    return await response.read(), response.status
            elif method == "POST":
                async with session.post(
                    url,
                    auth=auth,
                    headers=headers,
                    json=data,
                    ssl=False,
                    timeout=timeout,
                ) as response:
                    return await response.read(), response.status
            else:
                raise ValueError(f"Invalid HTTP method: {method}")
        except Exception as e:
            print_debug(
                f"Exception occurred: {str(e)}, at line {sys.exc_info()[-1].tb_lineno}"
            )
            if retry < retries - 1:  # if it's not the last retry attempt
                print_debug(f"Failed to fetch, retrying in {2 ** retry} seconds...")
                await asyncio.sleep(2**retry)
            else:
                print_debug(f"Error {e}")
                raise e


def error_queries(input, address_df):
    invalid_addresses = set()

    for region, data in input.items():
        for application, app in data.items():
            sources = set(app.get("source", []))
            destinations = set(app.get("destination", []))

            addresses = sources.union(destinations)
            invalid_addresses.update(
                addresses
                - set(address_df.loc[address_df["status"] == "VALID", "address"])
            )

    error_df = address_df[address_df["address"].isin(invalid_addresses)].copy()
    error_df["hostname"] = error_df["address"].apply(resolve_ip_to_domain)

    error_messages = error_df.apply(
        lambda row: f"Error occurred. Address: {row['address']}, Name: {row['hostname']}, Origin: {row['origin']} Status: {row['status']}, Description: {row['description']}",
        axis=1,
    )

    print("\nErrors:\n")
    print("\n".join(error_messages))

    return error_df


def filter_queries(input_df, address_df):
    if debug:
        print_debug("Calling: filter_queries")
        print_debug(f"Length of input_df: {len(input_df)}")
    try:
        # Flatten the dataframe to have a record for each address in the tuple
        address_df = address_df.explode("address")
        address_df = address_df.drop_duplicates(subset="address")
        # Convert address_df to a dictionary for faster lookup
        address_dict = address_df.set_index("address")[
            ["origin", "description", "status"]
        ].to_dict("index")

        queries = []
        for index, row in input_df.iterrows():
            region = row["region"]
            sources = row["sources"]
            destinations = row["destinations"]
            protocols = row["protocols"]
            dstPorts = row["dstPorts"]
            application = row["application"]
            action = row["action"]

            for source in sources:
                for destination in destinations:
                    if source != destination:
                        source_info = address_dict.get(source, {})
                        dest_info = address_dict.get(destination, {})

                        # Check if protocols is a range
                        if "-" in protocols and protocols != "0-255":
                            start, end = map(int, protocols.split("-"))
                            for proto in range(start, end + 1):
                                query = {
                                    "srcIp": source,
                                    "dstIp": destination,
                                    "ipProto": protocols
                                    if not re.match(r"\s+-\s+", protocols)
                                    else {},
                                    "dstPort": dstPorts
                                    if str(protocols) in ["6", "17"]
                                    else None,
                                    "source_origin": source_info.get("origin"),
                                    "source_description": source_info.get(
                                        "description"
                                    ),
                                    "source_status": source_info.get("status"),
                                    "dest_origin": dest_info.get("origin"),
                                    "dest_description": dest_info.get("description"),
                                    "dest_status": dest_info.get("status"),
                                    "region": region,
                                    "application": application,
                                    "AclAction": action,
                                }
                                queries.append(query)
                        else:
                            query = {
                                "region": region,
                                "application": application,
                                "srcIp": source,
                                "dstIp": destination,
                                "AclAction": action,
                                "dstPort": dstPorts
                                if protocols in ["6", "17"]
                                else None,
                                "source_origin": source_info.get("origin"),
                                "source_description": source_info.get("description"),
                                "source_status": source_info.get("status"),
                                "dest_origin": dest_info.get("origin"),
                                "dest_description": dest_info.get("description"),
                                "dest_status": dest_info.get("status"),
                            }
                            if protocols != "0-255":
                                query["ipProto"] = protocols
                            queries.append(query)
    except Exception as e:
        raise e
        # print_debug(f"An error occurred while creating address_dict1: {e}")

    if debug:
        for item in queries:
            print_debug(item)

    return queries


async def process_input(
    session, appserver, snapshot, input_df, batch_size, max_query, address_df=None
):
    if debug:
        print_debug("calling process_input")

    try:
        queries = filter_queries(input_df, address_df)
        # configure query limit
        total_queries = min(len(queries), max_query)
        dfs = []  # List to store individual dataframes

        # Iterate across queries for each region and application
        for region in input_df["region"].unique():
            for application in input_df[input_df["region"] == region][
                "application"
            ].unique():
                region_application_queries = [
                    query
                    for query in queries
                    if query["region"] == region and query["application"] == application
                ]
                total_queries = min(len(region_application_queries), max_query)
                print(
                    f"\nSearch Queries/Found for Region: {region}, Application: {application}: {total_queries}/{len(region_application_queries)}\n"
                )

                if total_queries > 0:
                    num_batches = math.ceil(total_queries / batch_size)
                    for i in tqdm(range(num_batches), desc="Path Search"):
                        start_index = i * batch_size
                        end_index = min((i + 1) * batch_size, total_queries)
                        batch_queries = [
                            {
                                k: v
                                for k, v in query.items()
                                if k in ["srcIp", "dstIp", "ipProto", "dstPort"]
                            }
                            for query in region_application_queries[
                                start_index:end_index
                            ]
                        ]
                        body = {"queries": batch_queries, **options}
                        if debug:
                            print_debug(batch_queries)
                        url = (
                            f"https://{appserver}/api/snapshots/{snapshot}/pathsBulkSeq"
                        )
                        try:
                            response_text, response_status = await fetch(
                                session,
                                url,
                                body,
                                method="POST",
                                username=username,
                                password=password,
                                headers=headers_seq,
                            )
                        except asyncio.TimeoutError:
                            print_debug(
                                "Request timed out. Skipping to next iteration."
                            )
                            continue
                        await asyncio.sleep(3)

                        parsed_data = []
                        # Check if the request was successful.
                        if response_status != 200:
                            print_debug(
                                f"Request failed with status code: {response_status}\n result: {response_text}\n body: {body}"
                            )
                            continue

                        lines = response_text.decode().split("\x1E")
                        parsed_data.extend(json.loads(line) for line in lines if line)
                        # Cleanup for dataframe import
                        try:
                            fix_data = check_info_paths(parsed_data)
                        except Exception as e:
                            print(f"Error occurred while checking info paths: {e}")
                            continue

                        paths_df = pd.json_normalize(
                            fix_data,
                            record_path=["info", "paths"],
                            meta=[
                                "dstIpLocationType",
                                "srcIpLocationType",
                                "pathCount",
                                "forwardHops",
                                "returnPathCount",
                                "returnHops",
                                "queryUrl",
                            ],
                            # errors="ignore",
                        )
                        logging.info("paths_df")
                        logging.warning(paths_df.iloc[0])

                        merged_df = pd.merge(
                            paths_df,
                            pd.DataFrame(
                                region_application_queries[start_index:end_index]
                            ),
                            left_index=True,
                            right_index=True,
                        )
                        merged_df.to_csv(
                            f"./cache/intent_{region}_{application}_{i}.csv",
                            index=False,
                        )
                        logging.info("merged_df")
                        logging.warning(merged_df.iloc[0])
                        dfs.append(merged_df)

        if len(dfs) > 0:
            return pd.concat(dfs, ignore_index=True)
    except Exception as e:
        print(f"An error occurred at line {sys.exc_info()[-1].tb_lineno}: {e}")
        raise e
        return None


def search_address(input):
    addresses = {
        a
        for region_value in input.values()
        for service_value in region_value.values()
        if isinstance(service_value, dict)
        for key, value in service_value.items()
        if key in ["source", "destination"]
        for a in value
    }
    return addresses


async def nqe_get_hosts_by_port(queryId, appserver, snapshot, device, port):
    async with aiohttp.ClientSession() as session:
        url = f"https://{appserver}/api/nqe?snapshotId={snapshot}"
        body = {
            "queryId": queryId,
            "queryOptions": {
                "columnFilters": [
                    {"columnName": "deviceName", "value": device},
                    {"columnName": "Interface", "value": port},
                ],
                "limit": 10000,
            },
        }
        try:
            response_text, response_status = await fetch(
                session,
                url,
                body,
                method="POST",
                username=username,
                password=password,
                headers=headers,
            )
            response_status.raise_for_status()
            response_json = json.loads(response_text)
            return response_json["items"]
        except asyncio.exceptions.TimeoutError:
            print_debug("Request timed out. Retrying...")
            return await nqe_get_hosts_by_port(
                queryId, appserver, snapshot, device, port
            )
        except requests.exceptions.HTTPError as err:
            if response_status in [401, 403]:
                print(
                    "Please set environment for FWD_USER, FWD_PASSWORD to your credentials"
                )
                sys.exit(1)
            else:
                print(f"Error: {err}")
                return None


async def nqe_get_hosts_by_port_2(session, appserver, snapshot):
    print(f"\nGathering Hosts Details...\n")
    url = f"https://{appserver}/api/nqe?snapshotId={snapshot}"
    offset = 0
    limit = 2000
    items = []
    while True:
        body = {
            "query": host_query,
            "queryOptions": {"offset": offset, "limit": limit},
        }
        try:
            response_text, response_status = await fetch(
                session,
                url,
                body,
                method="POST",
                username=username,
                password=password,
                headers=headers,
            )
        except asyncio.exceptions.TimeoutError:
            print_debug("Request timed out. Retrying...")
            continue
        if response_status == 200:
            response_json = json.loads(response_text)
            if len(response_json["items"]) == 0:
                break
            if debug:
                print_debug(f"Items:{len(response_json['items'])} Offset:{offset}")
            items.extend(response_json["items"])
            offset += limit
        else:
            print(f"Error: {response_status} {response_text}")
            continue
    print(f"Completed gathering hosts {len(items)}")
    df = pd.DataFrame(items)
    df.to_csv("./cache/hosts.csv")
    return df


async def search_subnet(session, appserver, snapshot, addresses):
    # Split the addresses into chunks for each coroutine
    num_coroutines = 6  # Change this to the number of coroutines you want
    address_chunks = [addresses[i::num_coroutines] for i in range(num_coroutines)]

    # Create a coroutine for each chunk of addresses
    coroutines = [
        search_subnet_chunk(session, appserver, snapshot, chunk)
        for chunk in address_chunks
    ]

    # Run the coroutines concurrently and gather the results
    results = await asyncio.gather(*coroutines)

    # Flatten the list of results
    result_list = [item for sublist in results for item in sublist]

    parsed_data = parse_subnets(result_list)
    df = pd.DataFrame(parsed_data)
    return df


async def search_subnet_chunk(session, appserver, snapshot, addresses):
    result_list = []  # List to store the response JSON for each address
    for address in tqdm(addresses, desc="Searching Subnets"):
        url = f"https://{appserver}/api/snapshots/{snapshot}/subnets"
        params = {"address": address, "minimal": "true"}
        try:
            response_text, response_status = await fetch(
                session,
                url,
                params,
                method="GET",
                username=username,
                password=password,
                headers=headers,
            )
        except asyncio.exceptions.TimeoutError:
            print_debug("Request timed out. Skipping to next iteration.")
            continue

        if response_status == 200:
            response_json = json.loads(response_text)
            response_json["address"] = address
            result_list.append(response_json)
        elif response_status == 403:
            print("Warning: Status 403. Please set FWD_USER and FWD_PASSWORD.")
            sys.exit(1)
        elif response_status == 400:
            raise Exception(f"Error: {response_text}")
        else:
            raise Exception(f"Error: {response_status}")

    return result_list


async def search_interface(session, appserver, snapshot, devices):
    if debug:
        print_debug("calling search_interface")
    df_list = []  # Initialize an empty list to store dataframes
    for device, interfaceName in tqdm(devices, desc="Searching Interfaces"):
        if interfaceName is not None:
            interface_update = toQuote(interfaceName)
            url = f"https://{appserver}/api/snapshots/{snapshot}/devices/{device}/interfaces/{interface_update}"
            try:
                response_text, response_status = await fetch(
                    session,
                    url,
                    data=None,
                    method="GET",
                    username=username,
                    password=password,
                    headers=headers,
                )
                response_json = json.loads(response_text)
                response_json["device"] = device
                if debug:
                    print_debug(response_json)
                df = pd.DataFrame(
                    [response_json],
                    columns=[
                        "name",
                        "type",
                        "description",
                        "ipAddresses",
                        "device",
                    ],
                )
                df = df.applymap(
                    lambda x: tuple(x) if isinstance(x, list) else x
                )  # Convert lists to tuples
                df_list.append(df)  # Add the dataframe to the list
            except asyncio.exceptions.TimeoutError:
                print_debug("Request timed out. Skipping to next iteration.")
                continue
            except KeyboardInterrupt:
                print_debug("Interrupted by user, raising exception...")
                raise
            except json.JSONDecodeError:
                print_debug(
                    f"Error: Could not decode the response into JSON. Response: {response_text}"
                )
                continue
            except Exception as e:
                print_debug(f"Error: An unexpected error occurred. {e}")
                print_debug(traceback.format_exc())
                continue
            if response_status == 403:
                print("Warning: Status 403. Please set FWD_USER and FWD_PASSWORD.")
                sys.exit(1)
            elif response_status != 200:
                print(
                    f"Warning: Unexpected status code {response_status}. Skipping to next iteration."
                )
                continue
    if df_list:
        return pd.concat(df_list).drop_duplicates()
    else:
        return pd.DataFrame(df_list)


def prepare_report(intent, hosts, app_df):
    if intent.empty:
        print("Intent is empty. Exiting early.")
        return
    forwarding_outcomes = addForwardingOutcomes(intent)

    report_df = return_firstlast_hop(forwarding_outcomes)

    for index, _ in tqdm(report_df.iterrows(), desc="Processing Data"):
        device = report_df.at[index, "lastHopDevice"]
        egressInterface = report_df.at[index, "lastHopEgressIntf"]
        dstIpLocationType = report_df.at[index, "dstIpLocationType"]
        forwardingOutcome = report_df.at[index, "forwardingOutcome"]
        securityOutcome = report_df.at[index, "securityOutcome"]
        dest_status = report_df.at[index, "dest_status"]
        behaviors = report_df.at[index, "ACLS"]
        outcomes = ["DELIVERED", "NOT_DELIVERED"]
        dstIp = report_df.at[index, "dstIp"]
        aclAction = report_df.at[index, "AclAction"]

        if (
            device
            and forwardingOutcome
            and egressInterface
            and not bool(re.match(r"^self\..*", egressInterface))
        ):
            host = hosts[
                (hosts["deviceName"] == device)
                & (hosts["Interface"] == egressInterface)
            ]
            if not host.empty:
                ipv4_regex = re.compile(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$")

                report_df.at[index, "hostAddress"] = ", ".join(
                    map(
                        str,
                        [
                            address
                            for address in set(host["Address"].values)
                            if ipv4_regex.match(address)
                        ],
                    )
                )
                report_df.at[index, "MacAddress"] = ", ".join(
                    map(str, set(host["MacAddress"].values))
                )
                report_df.at[index, "OUI"] = ", ".join(
                    map(str, set(host["OUI"].values))
                )

                report_df.at[index, "Disposition"] = getDisposition(
                    dstIp,
                    host["Address"].values,
                    egressInterface,
                    dstIpLocationType,
                    securityOutcome,
                    forwardingOutcome,
                    dest_status,
                )
                report_df.at[index, "Violation"] = (
                    False
                    if (
                        report_df.at[index, "securityOutcome"].lower() == "permitted"
                        and report_df.at[index, "AclAction"].lower() == "permit"
                    )
                    or (
                        report_df.at[index, "securityOutcome"].lower() == "denied"
                        and report_df.at[index, "AclAction"].lower() == "deny"
                    )
                    else True
                )
                report_df.at[index, "Diagnostic"] = getDiagnostic(
                    dstIp,
                    host["Address"].values,
                    egressInterface,
                    dstIpLocationType,
                    securityOutcome,
                    forwardingOutcome,
                    dest_status,
                    behaviors,
                    app_df,
                    report_df.at[index, "Violation"],
                    aclAction
                )

                # report_df.at[index, "hostInterface"] = host["Interface"].values[0]; this would the same, not sure if we should check
                logging.info(
                    f"Updated host details for device: {device} and interface: {egressInterface}"
                )
            else:
                logging.warning(
                    f"No host details found for device: {device} and interface: {egressInterface}"
                )
                report_df.at[index, "hostAddress"] = None
                report_df.at[index, "MacAddress"] = None
                report_df.at[index, "OUI"] = None
                report_df.at[index, "Disposition"] = getDisposition(
                    dstIp,
                    host["Address"].values,
                    egressInterface,
                    dstIpLocationType,
                    securityOutcome,
                    forwardingOutcome,
                    dest_status,
                )
                report_df.at[index, "Violation"] = (
                    False
                    if (
                        report_df.at[index, "securityOutcome"].lower() == "permitted"
                        and report_df.at[index, "AclAction"].lower() == "permit"
                    )
                    or (
                        report_df.at[index, "securityOutcome"].lower() == "denied"
                        and report_df.at[index, "AclAction"].lower() == "deny"
                    )
                    else True
                )
                report_df.at[index, "Diagnostic"] = getDiagnostic(
                    dstIp,
                    None,
                    egressInterface,
                    dstIpLocationType,
                    securityOutcome,
                    forwardingOutcome,
                    dest_status,
                    behaviors,
                    app_df,
                    report_df.at[index, "Violation"],
                    aclAction
                )

        else:
            report_df.at[index, "hostAddress"] = None
            report_df.at[index, "MacAddress"] = None
            report_df.at[index, "OUI"] = None
            # report_df.at[index, "Disposition"] = False
            report_df.at[index, "Disposition"] = getDisposition(
                dstIp,
                None,
                egressInterface,
                dstIpLocationType,
                securityOutcome,
                forwardingOutcome,
                dest_status,
            )
            report_df.at[index, "Violation"] = (
                False
                if (
                    report_df.at[index, "securityOutcome"].lower() == "permitted"
                    and report_df.at[index, "AclAction"].lower() == "permit"
                )
                or (
                    report_df.at[index, "securityOutcome"].lower() == "denied"
                    and report_df.at[index, "AclAction"].lower() == "deny"
                )
                else True
            )
            report_df.at[index, "Diagnostic"] = getDiagnostic(
                    dstIp,
                    None,
                    egressInterface,
                    dstIpLocationType,
                    securityOutcome,
                    forwardingOutcome,
                    dest_status,
                    behaviors,
                    app_df,
                    report_df.at[index, "Violation"],
                    aclAction
                )
            # report_df.at[index, "hostInterface"] = None
            logging.warning(
                f"No device or interface details found for device: {device} and interface: {egressInterface}"
            )

    return report_df

def prepare_report2(intent, hosts, app_df):
    if intent.empty:
        print("Intent is empty. Exiting early.")
        return
    forwarding_outcomes = addForwardingOutcomes(intent)
    report_df = return_firstlast_hop(forwarding_outcomes)

    for index, _ in tqdm(report_df.iterrows(), desc="Processing Data"):
        row = report_df.loc[index]
        device, egressInterface = row["lastHopDevice"], row["lastHopEgressIntf"]
        dstIpLocationType, forwardingOutcome = row["dstIpLocationType"], row["forwardingOutcome"]
        securityOutcome, dest_status = row["securityOutcome"], row["dest_status"]
        behaviors, dstIp, aclAction = row["ACLS"], row["dstIp"], row["AclAction"]
    

        if device and forwardingOutcome and egressInterface and not bool(re.match(r"^self\..*", egressInterface)):
            host = hosts[(hosts["deviceName"] == device) & (hosts["Interface"] == egressInterface)]
            hostAddress = ", ".join(
                    map(
                        str,
                        [
                            address
                            for address in set(host["Address"].values)
                            if ipv4_regex.match(address)
                        ],
                    )
                )
            if not host.empty:
                ipv4_regex = re.compile(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$")
                report_df.at[index, "hostAddress"], report_df.at[index, "MacAddress"], report_df.at[index, "OUI"] = get_host_details(host, ipv4_regex)
                report_df.at[index, "Disposition"], report_df.at[index, "Violation"] = getDisposition(row, host, egressInterface, dstIpLocationType, securityOutcome, forwardingOutcome, dest_status)
                logging.info(f"Updated host details for device: {device} and interface: {egressInterface}")
            else:
                logging.warning(f"No host details found for device: {device} and interface: {egressInterface}")
                report_df.at[index, "hostAddress"], report_df.at[index, "MacAddress"], report_df.at[index, "OUI"] = None, None, None
                report_df.at[index, "Disposition"] = getDisposition(row, host, egressInterface, dstIpLocationType, securityOutcome, forwardingOutcome, dest_status)
                report_df.at[index, "Diagnostic"] = getDiagnostic(row, host, egressInterface, dstIpLocationType, securityOutcome, forwardingOutcome, dest_status, behaviors, app_df)
        else:
            report_df.at[index, "hostAddress"], report_df.at[index, "MacAddress"], report_df.at[index, "OUI"] = None, None, None
            report_df.at[index, "Disposition"] = getDisposition(row, host, egressInterface, dstIpLocationType, securityOutcome, forwardingOutcome, dest_status)
            report_df.at[index, "Diagnostic"] = getDiagnostic(row, host, egressInterface, dstIpLocationType, securityOutcome, forwardingOutcome, dest_status, behaviors, app_df)
            logging.warning(f"No device or interface details found for device: {device} and interface: {egressInterface}")
    return report_df


def generate_report(snapshot, report_df, with_diag=False):
    if debug:
        print_debug("calling generate_report")
    report = f"intent-gen-{snapshot}.csv"

    columns_to_display = [
        "region",
        "application",
        "srcIp",
        "dstIp",
        "ipProto",
        "dstPort",
        "source_status",
        "dest_status",
        "forwardingOutcome",
        "securityOutcome",
        "Violation",
        "Disposition",
        "Diagnostic",
        "AclAction",
        "srcIpLocationType",
        "dstIpLocationType",
        "firstHopDevice",
        "lastHopDevice",
        "lastHopEgressIntf",
        "hostAddress",
        "MacAddress",
        "OUI",
        "pathCount",
        "forwardHops",
        "returnPathCount",
        "returnHops",
        "ACLS",
    ]

    # Excel has a max row limit of 1048576
    # report_df = report_df.head(1048575)

    try:
        if with_diag:
            report_df[
                columns_to_display
                + [
                    "queryUrl",
                    "forwardDescription",
                    "forwardRemedy",
                    "securityDescription",
                    "securityRemedy",
                ]
            ].to_csv(report, index=False)
        else:
            report_df[
                columns_to_display
                + [
                    "queryUrl",
                ]
            ].to_csv(report, index=False)

        print(f"Report Created: {report}")
    except Exception as e:
        print(f"Error occurred while writing to CSV: {e}")
        raise
    # update_font(report)


async def handler(
    appserver, snapshot, addresses, app_df, batchsize, max_query, with_diag
):
    async with aiohttp.ClientSession() as session:
        try:
            address_df = await search_subnet(session, appserver, snapshot, addresses)

            address_df.to_csv(f"./cache/subnets.csv", index=False)
            hosts = await nqe_get_hosts_by_port_2(session, appserver, snapshot)

            start_time = datetime.datetime.now()
            print(f"\n\nStart processor time: {start_time}")
            logging.info(f"Start processor time: {start_time}")

            results = await process_input(
                session,
                appserver,
                snapshot,
                app_df,
                batchsize,
                max_query,
                address_df,
            )
            intent = pd.DataFrame(results)

            print(f"Collection End: {datetime.datetime.now()}")
            logging.info(f"Collection End: {datetime.datetime.now()}")

            report_df = prepare_report(intent, hosts, app_df)

            addresses = list(
                set([tuple(x) for x in report_df[["srcIp", "dstIp"]].values.tolist()])
            )
            combined_list = list(
                set(
                    zip(
                        report_df["lastHopDevice"].tolist(),
                        report_df["lastHopEgressIntf"].tolist(),
                    )
                )
            )

            last_hop_address_lookup_df = await search_interface(
                session, appserver, snapshot, combined_list
            )
            last_hop_address_lookup_df.to_csv(f"./cache/lastHopDevice.csv", index=False)
            generate_report(snapshot, report_df, with_diag)

        except asyncio.TimeoutError:
            print_debug(
                "Operation timed out. Recovering from the persisted dataframe..."
            )
            if os.path.exists("./cache/hosts.csv") and glob.glob(
                "./cache/intent_*.csv"
            ):
                hosts = pd.read_csv("./cache/hosts.csv")
                csv_files = glob.glob("./cache/intent_*.csv")
                intent = pd.concat(
                    [pd.read_csv(f) for f in csv_files], ignore_index=True
                )
            else:
                print("Exiting Early")
                return
        except aiohttp.ClientOSError:
            print_debug(
                "Operation timed out. Recovering from the persisted dataframe..."
            )
            if os.path.exists("./cache/hosts.csv") and glob.glob(
                "./cache/intent_*.csv"
            ):
                csv_files = glob.glob("./cache/intent_*.csv")
                hosts = pd.read_csv("./cache/hosts.csv")
                intent = pd.concat(
                    [pd.read_csv(f) for f in csv_files], ignore_index=True
                )
            else:
                print("Exiting Early")
                return


def from_import(
    appserver, snapshot, infile, batchsize, limit, max_query, retries, with_diag
):
    print(f"Setting batch size: {batchsize}")
    print(f"Setting application limit: {limit}")
    print(f"Setting max queries: {max_query}\n")

    test_communication(appserver)

    if debug:
        pd.set_option("display.max_rows", None)  # Show all rows

    with open(infile) as file:
        data = json.load(file)

    app_df = flatten_input(data)

    addresses = []
    for row in app_df.itertuples():
        addresses.extend(row.sources)
        addresses.extend(row.destinations)
        addresses = list(set(addresses))  # remove duplicates

        app_df.sort_values(by="application")

        # Fix list
        # for column in app_df.columns:
        #     if app_df[column].apply(lambda x: isinstance(x, list)).any():
        #         app_df[column] = app_df[column].apply(tuple)

        app_df = app_df.applymap(lambda x: tuple(x) if isinstance(x, list) else x)

        app_df.drop_duplicates(inplace=True)
        print(f"APP Entries Found: {len(app_df)}\n")
        app_df.to_csv(f"./cache/acls.csv", index=False)

        # add limiter for testing
        if limit:
            app_df = app_df.head(limit)

        asyncio.run(
            handler(
                appserver,
                snapshot,
                addresses,
                app_df,
                batchsize,
                max_query,
                with_diag,
            )
        )


def from_acls(
    appserver, snapshot, batchsize, limit, max_query, retries, with_diag=False
):
    print(f"Setting batch size: {batchsize}")
    print(f"Setting application limit: {limit}")
    print(f"Setting max querys: {max_query}\n")
    # if debug:
    #     pd.set_option("display.max_rows", None)  # Show all rows
    try:
        test_communication(appserver)
        # Retrieve all possible ACLs where a source or destination is locatable in the model

        print("Retrieving ACL Entries... This can take a while")
        data = nqe_get_hosts_from_acl(acl_query, appserver, snapshot)
        print(f"ACL Entries retrieved successfully.")
        app_df = pd.DataFrame(data)
        if len(app_df) == 0:
            print("No ACL names found")
            return

        # Add regions and fixup ports
        app_df.sort_values(by="application")
        app_df["region"] = "Default"
        app_df["dstPorts"] = app_df["dstPorts"].apply(parse_start_end)
        app_df["protocols"] = app_df["protocols"].apply(parse_start_end)

        #  # Fix list
        # for column in app_df.columns:
        #     if app_df[column].apply(lambda x: isinstance(x, list)).any():
        #         app_df[column] = app_df[column].apply(tuple)
        app_df = app_df.applymap(lambda x: tuple(x) if isinstance(x, list) else x)

        app_df.drop_duplicates(inplace=True)
        app_df.to_csv(f"./cache/acls.csv", index=False)
        # add limiter for testing
        if limit:
            app_df = app_df.head(limit)

        addresses = []
        for row in app_df.itertuples():
            addresses.extend(row.sources)
            addresses.extend(row.destinations)
        addresses = pd.unique(pd.concat([app_df["sources"], app_df["destinations"]]))
        num_addresses = len(addresses)
        num_data = len(data)

        print(f"ACL Entries Found: {num_data}, Addresses: {num_addresses}\n")

        asyncio.run(
            handler(
                appserver,
                snapshot,
                addresses,
                app_df,
                batchsize,
                max_query,
                with_diag,
            )
        )

    except Exception as e:
        print_debug(f"An error occurred at line {sys.exc_info()[-1].tb_lineno}: {e}")
        print(traceback.format_exc())
        logging.error(f"An error occurred: {e}")
        logging.error(traceback.format_exc())


def check(appserver, snapshot, infile, csv):
    report = f"errored-devices-{snapshot}.csv"

    test_communication(appserver)

    if debug:
        pd.set_option("display.max_rows", None)  # Show all rows

    with open(infile) as file:
        data = json.load(file)

    app_df = flatten_input(data)
    try:
        address_df = asyncio.run(search_subnet(appserver, snapshot, app_df))
    except Exception as e:
        print(f"Error occurred while searching subnet: {e}")
        return

    try:
        errored_devices = error_queries(data, address_df)
    except Exception as e:
        print(f"Error occurred while querying errors: {e}")
        return

    if csv:
        errored_devices.loc[:, ["address", "hostname"]].to_csv(report, index=False)


def main():
    arguments = docopt(__doc__)
    infile = arguments["<input>"]
    appserver = arguments["<appserver>"]
    snapshot = arguments["<snapshot>"]
    batchsize = int(arguments["--batch"])
    limit = int(arguments["--limit"])
    max_query = int(arguments["--max"])
    with_diag = arguments["--withdiag"]
    global debug
    debug = arguments["--debug"]
    print(f"Debug: {debug}")
    retries = 3
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    if not os.path.exists("./logs"):
        os.makedirs("./logs")
    logging.basicConfig(
        filename=f"./logs/fwd-intent-gen-{timestamp}.log", level=logging.INFO
    )

    if not os.path.exists("./cache"):
        os.makedirs("./cache")

    # Check for existing intent*.csv files in ./cache directory
    csv_files = glob.glob("./cache/*.csv")
    if csv_files:
        print("Found existing intent*.csv files in ./cache directory.")
        purge = input("Do you want to purge these results? (yes)/no: ")
        if purge.lower() == "yes" or purge == "":
            for file in tqdm(csv_files, desc="Purging cache\n"):
                os.remove(file)
        else:
            csv_files = glob.glob("./cache/intent_*.csv")
            intent = pd.concat([pd.read_csv(f) for f in csv_files], ignore_index=True)
            intent["hops"] = intent["hops"].apply(
                lambda x: json.loads(x.replace("'", '"'))
            )

            print(f"Total rows in intent: {len(intent)}")

            hosts = pd.read_csv("./cache/hosts.csv")
            app_df = pd.read_csv("./cache/acls.csv")

            report_df = prepare_report(intent, hosts, app_df)

            generate_report(snapshot, report_df, with_diag)
            return

    if arguments["from_import"]:
        print("Running: from_import")

        from_import(
            appserver, snapshot, infile, batchsize, limit, max_query, retries, with_diag
        )

    elif arguments["check"]:
        print("Running: Check")
        infile = arguments["<input>"]
        appserver = arguments["<appserver>"]
        snapshot = arguments["<snapshot>"]
        csv = arguments["--csv"]

        check(appserver, snapshot, infile, csv)

    elif arguments["from_acls"]:
        print("Running: from_acls")
        appserver = arguments["<appserver>"]
        snapshot = arguments["<snapshot>"]
        batchsize = int(arguments["--batch"])
        limit = int(arguments["--limit"])
        max_query = int(arguments["--max"])
        with_diag = arguments["--withdiag"]
        from_acls(appserver, snapshot, batchsize, limit, max_query, retries, with_diag)

    else:
        print(
            "Invalid command. Please refer to the usage message for available commands."
        )


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nInterrupted by user. Exiting...")
        sys.exit(0)
